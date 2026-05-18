import boto3
from botocore.exceptions import ClientError, ProfileNotFound
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
from typing import List, Dict, Any, Optional
import configparser
import os

# --- Configuration ---
# The base name for the output file.
OUTPUT_FILENAME_BASE = 'aws_ec2_inventory_linux'

def get_available_profiles() -> List[str]:
    """
    Reads the AWS credentials and config files to find all available profile names.
    """
    profiles = set()
    
    # Standard paths for AWS configuration files
    credentials_path = os.path.expanduser('~/.aws/credentials')
    config_path = os.path.expanduser('~/.aws/config')
    
    parser = configparser.ConfigParser()
    
    if os.path.exists(credentials_path):
        parser.read(credentials_path)
        profiles.update(parser.sections())
        
    if os.path.exists(config_path):
        parser.read(config_path)
        for section in parser.sections():
            if section.startswith('profile '):
                profiles.add(section.replace('profile ', '').strip())
            else:
                profiles.add(section)

    if 'default' not in profiles and 'default' in parser:
        profiles.add('default')
        
    return sorted(list(profiles))

def _get_ssm_statuses(session: boto3.Session, region: str, instance_ids: List[str]) -> Dict[str, str]:
    """Helper function to fetch SSM agent statuses for a list of instance IDs."""
    if not instance_ids:
        return {}

    ssm_status_map = {}
    try:
        ssm_client = session.client('ssm', region_name=region)
        
        for i in range(0, len(instance_ids), 50):
            chunk = instance_ids[i:i + 50]
            paginator = ssm_client.get_paginator('describe_instance_information')
            page_iterator = paginator.paginate(
                Filters=[{'Key': 'InstanceIds', 'Values': chunk}]
            )
            for page in page_iterator:
                for info in page.get('InstanceInformationList', []):
                    ssm_status_map[info['InstanceId']] = info.get('PingStatus', 'Unknown')
    except ClientError as e:
        print(f"Could not get SSM statuses in region {region}. Check IAM permissions. Error: {e}")
        return {}
        
    return ssm_status_map

def _get_ami_details(session: boto3.Session, region: str, ami_ids: List[str]) -> Dict[str, str]:
    """Helper function to fetch AMI names for a list of AMI IDs."""
    if not ami_ids:
        return {}

    ami_details_map = {}
    try:
        ec2_client = session.client('ec2', region_name=region)
        images = ec2_client.describe_images(ImageIds=ami_ids)
        for image in images.get('Images', []):
            ami_details_map[image['ImageId']] = image.get('Name', 'N/A')
    except ClientError as e:
        print(f"Could not get AMI details in region {region}. Check permissions for ec2:DescribeImages. Error: {e}")
        return {ami_id: 'Permission Error' for ami_id in ami_ids}
        
    return ami_details_map

def parse_os_version(ami_name: str) -> str:
    """Parses a raw AMI name to find a simple, readable OS version."""
    if not ami_name or ami_name in ['N/A', 'AMI Details Not Found', 'Permission Error']:
        return ami_name

    # Make the search case-insensitive
    lower_name = ami_name.lower()

    # --- RHEL Checks ---
    if 'rhel-9' in lower_name or 'rhel9' in lower_name:
        return 'RHEL 9'
    if 'rhel-8' in lower_name or 'rhel8' in lower_name:
        return 'RHEL 8'
    if 'rhel-7' in lower_name or 'rhel7' in lower_name:
        return 'RHEL 7'

    # --- Other Common OS Checks (Optional but recommended) ---
    if 'ubuntu-22.04' in lower_name or 'jammy' in lower_name:
        return 'Ubuntu 22.04'
    if 'ubuntu-20.04' in lower_name or 'focal' in lower_name:
        return 'Ubuntu 20.04'
    if 'ubuntu' in lower_name:
        return 'Ubuntu (Other)'
    if 'amzn2' in lower_name or 'amazon linux 2' in lower_name:
        return 'Amazon Linux 2'
    if 'suse' in lower_name:
        return 'SUSE'
        
    # If no specific version is found, return the original name
    return ami_name

def get_all_ec2_instances(profile_name: str, region_input: str) -> Optional[List[Dict[str, Any]]]:
    """
    Connects to AWS using a specific profile and fetches details of all EC2 instances.
    """
    print(f"\n--- Processing Profile: {profile_name} ---")
    print(f"Attempting to connect to AWS with profile: '{profile_name}'...")
    try:
        session = boto3.Session(profile_name=profile_name)
    except ProfileNotFound:
        print(f"Error: The AWS profile '{profile_name}' was not found.")
        return None
    except Exception as e:
        print(f"An unexpected error occurred during session creation: {e}")
        return None

    all_instances = []
    regions_to_scan = []

    if region_input.lower() == 'all':
        print("Finding all available EC2 regions...")
        base_ec2_client = session.client('ec2', region_name='us-east-1')
        try:
            response = base_ec2_client.describe_regions()
            regions_to_scan = [region['RegionName'] for region in response['Regions']]
            print(f"Found {len(regions_to_scan)} regions to scan.")
        except ClientError as e:
            print(f"Error fetching AWS regions: {e}")
            return None
    else:
        regions_to_scan.append(region_input)

    for region in regions_to_scan:
        print(f"\n--- Scanning Region: {region} ---")
        try:
            ec2_client = session.client('ec2', region_name=region)
            paginator = ec2_client.get_paginator('describe_instances')
            page_iterator = paginator.paginate(
                Filters=[{'Name': 'instance-state-name', 'Values': ['pending', 'running', 'stopping', 'stopped']}]
            )
            
            regional_instances = []
            for page in page_iterator:
                for reservation in page['Reservations']:
                    for instance in reservation['Instances']:
                        instance['Region'] = region
                        regional_instances.append(instance)
            
            if not regional_instances:
                continue

            print(f"Found {len(regional_instances)} instances in {region}. Checking SSM status & AMI details...")
            
            instance_ids_in_region = [inst['InstanceId'] for inst in regional_instances]
            ssm_statuses = _get_ssm_statuses(session, region, instance_ids_in_region)
            
            unique_ami_ids = list(set(inst['ImageId'] for inst in regional_instances))
            ami_details = _get_ami_details(session, region, unique_ami_ids)

            for instance in regional_instances:
                instance_id = instance['InstanceId']
                ami_id = instance['ImageId']
                instance['SSMAgentStatus'] = ssm_statuses.get(instance_id, 'Not Managed')
                instance['OSVersionFromAMI'] = ami_details.get(ami_id, 'AMI Details Not Found')

            all_instances.extend(regional_instances)

        except ClientError as e:
            if e.response['Error']['Code'] in ['AuthFailure', 'UnauthorizedOperation']:
                print(f"Could not access region '{region}'. It might not be enabled for your account. Skipping.")
            else:
                print(f"An unexpected error occurred in region {region}: {e}")
            continue

    print(f"\nTotal instances found for profile {profile_name}: {len(all_instances)}.")
    return all_instances

def write_instances_to_sheet(workbook: openpyxl.Workbook, all_instances: List[Dict[str, Any]], profile_name: str) -> None:
    """
    Writes the collected instance data to a new sheet within the provided workbook.
    """
    if not all_instances:
        print(f"No instance data to write for profile '{profile_name}'.")
        return

    sheet = workbook.create_sheet(title=profile_name)

    headers = [
        'Instance Name', 'Instance ID', 'Private IPv4', 'Instance State', 
        'Instance Type', 'Region', 'Availability Zone', 'Platform', 'SSM Agent Status',
        'OS Version (Parsed)'
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    print(f"\nFiltering for Linux instances and writing data to sheet '{sheet.title}'...")
    linux_instances_written = 0
    for instance in all_instances:
        platform_details = instance.get('PlatformDetails', 'Linux/UNIX')

        if 'Windows' in platform_details:
            continue

        tags = {tag['Key']: tag['Value'] for tag in instance.get('Tags', [])}
        
        # Get the raw AMI name
        raw_ami_name = instance.get('OSVersionFromAMI', 'N/A')
        # Parse it using our new function
        simple_os_name = parse_os_version(raw_ami_name)
        
        row = [
            tags.get('Name', 'N/A'),
            instance.get('InstanceId', 'N/A'),
            instance.get('PrivateIpAddress', 'N/A'),
            instance.get('State', {}).get('Name', 'N/A'),
            instance.get('InstanceType', 'N/A'),
            instance.get('Region', 'N/A'),
            instance.get('Placement', {}).get('AvailabilityZone', 'N/A'),
            platform_details,
            instance.get('SSMAgentStatus', 'N/A'),
            simple_os_name
        ]
        sheet.append(row)
        linux_instances_written += 1
    
    print(f"Wrote {linux_instances_written} Linux instances to the sheet.")
    
def main() -> None:
    """
    Main function to orchestrate fetching data and writing the report for multiple accounts.
    """
    profiles_input = input("Enter AWS profile names (comma-separated), or 'all' to scan all profiles: ")
    region_name = input("Enter an AWS region (e.g., us-west-2) or 'all' to scan all regions: ")

    profile_names = []
    if profiles_input.lower() == 'all':
        print("Finding all available AWS profiles...")
        profile_names = get_available_profiles()
        if not profile_names:
            print("No AWS profiles found in your credentials or config files.")
            return
        print(f"Found profiles: {', '.join(profile_names)}")
    else:
        profile_names = [p.strip() for p in profiles_input.split(',')]

    if not profile_names or not profiles_input:
        print("At least one profile name is required. Please run the script again.")
        return
        
    if not region_name:
        print("Region cannot be empty. Please run the script again.")
        return

    output_file = f"{OUTPUT_FILENAME_BASE}.xlsx"

    if os.path.exists(output_file):
        try:
            os.remove(output_file)
            print(f"Removed old report file: {output_file}")
        except OSError as e:
            print(f"Error removing old report file: {e}")
            return

    workbook = openpyxl.Workbook()
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    for profile_name in profile_names:
        instances = get_all_ec2_instances(profile_name, region_name)
    
        if instances:
            write_instances_to_sheet(workbook, instances, profile_name)

    if len(workbook.sheetnames) > 0:
        try:
            workbook.save(output_file)
            print(f"\nSuccess! New Excel report created: '{output_file}'")
        except Exception as e:
            print(f"Error saving the Excel file: {e}")
    else:
        print("\nNo instances found across all profiles. The Excel file was not created.")

if __name__ == '__main__':
    main()