import boto3
from botocore.exceptions import ClientError, ProfileNotFound
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any, Optional
import configparser
import os
import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed

# --- Configuration ---
OUTPUT_FILENAME_BASE = 'aws_ec2_inventory_enhanced'

def get_available_profiles() -> List[str]:
    """Reads the AWS credentials and config files to find all available profile names."""
    profiles = set()
    credentials_path = os.path.expanduser('~/.aws/credentials')
    config_path = os.path.expanduser('~/.aws/config')
    parser = configparser.ConfigParser()
    
    if os.path.exists(credentials_path):
        parser.read(credentials_path)
        profiles.update(parser.sections())
        
    if os.path.exists(config_path):
        parser.read(config_path)
        for section in parser.sections():
            if section.startswith('profile '):
                profiles.add(section.replace('profile ', '').strip())
            else:
                profiles.add(section)

    if 'default' not in profiles and 'default' in parser:
        profiles.add('default')
        
    return sorted(list(profiles))

def _get_ssm_statuses(session: boto3.Session, region: str, instance_ids: List[str]) -> Dict[str, str]:
    """Helper function to fetch SSM agent statuses for a list of instance IDs."""
    if not instance_ids:
        return {}

    ssm_status_map = {}
    try:
        ssm_client = session.client('ssm', region_name=region)
        for i in range(0, len(instance_ids), 50):
            chunk = instance_ids[i:i + 50]
            paginator = ssm_client.get_paginator('describe_instance_information')
            page_iterator = paginator.paginate(Filters=[{'Key': 'InstanceIds', 'Values': chunk}])
            for page in page_iterator:
                for info in page.get('InstanceInformationList', []):
                    ssm_status_map[info['InstanceId']] = info.get('PingStatus', 'Unknown')
    except ClientError as e:
        print(f"Could not get SSM statuses in region {region}. Error: {e}")
        return {}
        
    return ssm_status_map

def scan_region(session: boto3.Session, region: str) -> List[Dict[str, Any]]:
    """Scans a single region for EC2 instances and their SSM status."""
    try:
        ec2_client = session.client('ec2', region_name=region)
        paginator = ec2_client.get_paginator('describe_instances')
        page_iterator = paginator.paginate(
            Filters=[{'Name': 'instance-state-name', 'Values': ['pending', 'running', 'stopping', 'stopped']}]
        )
        
        regional_instances = []
        for page in page_iterator:
            for reservation in page['Reservations']:
                for instance in reservation['Instances']:
                    instance['Region'] = region
                    regional_instances.append(instance)
        
        if not regional_instances:
            return []

        instance_ids = [inst['InstanceId'] for inst in regional_instances]
        ssm_statuses = _get_ssm_statuses(session, region, instance_ids)

        for instance in regional_instances:
            instance['SSMAgentStatus'] = ssm_statuses.get(instance['InstanceId'], 'Not Managed')

        return regional_instances

    except ClientError as e:
        if e.response['Error']['Code'] not in ['AuthFailure', 'UnauthorizedOperation']:
            print(f"Error scanning region {region}: {e}")
        return []

def get_all_ec2_instances(profile_name: str, region_input: str) -> Optional[List[Dict[str, Any]]]:
    """Fetches details of all EC2 instances using parallel regional scans."""
    print(f"\n--- Processing Profile: {profile_name} ---")
    try:
        session = boto3.Session(profile_name=profile_name)
    except ProfileNotFound:
        print(f"Error: Profile '{profile_name}' not found.")
        return None

    regions_to_scan = []
    if region_input.lower() == 'all':
        base_ec2_client = session.client('ec2', region_name='us-east-1')
        try:
            response = base_ec2_client.describe_regions()
            regions_to_scan = [r['RegionName'] for r in response['Regions']]
        except ClientError as e:
            print(f"Error fetching regions: {e}")
            return None
    else:
        regions_to_scan.append(region_input)

    all_instances = []
    print(f"Scanning {len(regions_to_scan)} regions in parallel...")
    with ThreadPoolExecutor(max_workers=10) as executor:
        future_to_region = {executor.submit(scan_region, session, region): region for region in regions_to_scan}
        for future in as_completed(future_to_region):
            all_instances.extend(future.result())

    print(f"Total instances found for profile {profile_name}: {len(all_instances)}.")
    return all_instances

def write_instances_to_sheet(workbook: openpyxl.Workbook, all_instances: List[Dict[str, Any]], profile_name: str) -> None:
    """Writes instance data to a new sheet with enhanced formatting and metadata."""
    if not all_instances:
        return

    sheet = workbook.create_sheet(title=profile_name[:31]) # Excel limit
    headers = [
        'Instance Name', 'Instance ID', 'Private IP', 'Public IP', 'State', 
        'Type', 'Region', 'AZ', 'VPC ID', 'Subnet ID', 'Launch Time', 'Platform', 'SSM Status'
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
    
    sheet.freeze_panes = 'A2'

    linux_count = 0
    for instance in all_instances:
        platform = instance.get('PlatformDetails', 'Linux/UNIX')
        if 'Windows' in platform: continue

        tags = {tag['Key']: tag['Value'] for tag in instance.get('Tags', [])}
        launch_time = instance.get('LaunchTime')
        if launch_time:
            launch_time = launch_time.strftime('%Y-%m-%d %H:%M:%S')

        row = [
            tags.get('Name', 'N/A'),
            instance.get('InstanceId', 'N/A'),
            instance.get('PrivateIpAddress', 'N/A'),
            instance.get('PublicIpAddress', 'N/A'),
            instance.get('State', {}).get('Name', 'N/A'),
            instance.get('InstanceType', 'N/A'),
            instance.get('Region', 'N/A'),
            instance.get('Placement', {}).get('AvailabilityZone', 'N/A'),
            instance.get('VpcId', 'N/A'),
            instance.get('SubnetId', 'N/A'),
            launch_time or 'N/A',
            platform,
            instance.get('SSMAgentStatus', 'N/A')
        ]
        sheet.append(row)
        linux_count += 1
    
    # Auto-filter and Auto-width
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    print(f"Wrote {linux_count} Linux instances to sheet '{sheet.title}'.")

def main():
    parser = argparse.ArgumentParser(description="AWS EC2 Inventory Script (Enhanced)")
    parser.add_argument("--profiles", help="Comma-separated profiles or 'all'", default=None)
    parser.add_argument("--regions", help="Comma-separated regions or 'all'", default='all')
    args = parser.parse_args()

    profiles_input = args.profiles or input("Enter AWS profile names (comma-separated), or 'all': ")
    regions_input = args.regions

    if profiles_input.lower() == 'all':
        profile_names = get_available_profiles()
    else:
        profile_names = [p.strip() for p in profiles_input.split(',')]

    if not profile_names:
        print("No profiles specified.")
        return

    output_file = f"{OUTPUT_FILENAME_BASE}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    workbook = openpyxl.Workbook()
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    for profile_name in profile_names:
        instances = get_all_ec2_instances(profile_name, regions_input)
        if instances:
            write_instances_to_sheet(workbook, instances, profile_name)

    if len(workbook.sheetnames) > 0:
        workbook.save(output_file)
        print(f"\nSuccess! Enhanced report created: '{output_file}'")
    else:
        print("\nNo instances found.")

if __name__ == '__main__':
    main()
